#!/usr/bin/env python3
"""
Sistem Absensi Karyawan - Pro Edition
Update Otomatis ke File Excel Hasil Akhir
Mendukung format input: CSV dan XLSX
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from datetime import datetime
import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
import calendar


class SistemAbsensiPro:
    """Aplikasi GUI untuk pengolahan data absensi dengan output ke hasil-akhir.xlsx"""
    
    # Hari libur nasional Indonesia 2026
    INDONESIAN_HOLIDAYS = {
        '2026': {
            '01-01': 'Tahun Baru 2026',
            '03-23': 'Isra Miraj',
            '03-31': 'Hari Suci Nyepi',
            '04-03': 'Wafat Yesus Kristus',
            '04-05': 'Paskah',
            '05-01': 'Hari Buruh',
            '05-04': 'Kenaikan Yesus Kristus',
            '05-14': 'Hari Raya Waisak',
            '06-01': 'Hari Lahir Pancasila',
            '06-17': 'Idul Fitri',
            '06-18': 'Idul Fitri',
            '08-17': 'Hari Kemerdekaan RI',
            '08-24': 'Idul Adha',
            '09-14': 'Tahun Baru Islam',
            '11-23': 'Maulid Nabi Muhammad SAW',
            '12-25': 'Hari Raya Natal',
            '12-26': 'Cuti Bersama Natal',
        }
    }
    
    def __init__(self, root):
        self.root = root
        self.root.title("Sistem Absensi Pro - Excel Updater v2.0")
        self.root.geometry("1000x800")
        self.root.configure(bg='#ecf0f1')
        
        # Data variables
        self.df_source = None
        self.employee_list = []
        self.processed_data = None
        self.available_periods = []  # List of (month, year) tuples
        
        # UI Variables
        self.input_file_path = tk.StringVar()
        self.selected_employee = tk.StringVar()
        self.selected_month = tk.IntVar(value=datetime.now().month)
        self.selected_year = tk.IntVar(value=datetime.now().year)
        self.output_file = tk.StringVar(value="hasil-akhir.xlsx")
        
        # Setup UI
        self.setup_ui()
        
    def setup_ui(self):
        """Setup semua komponen UI dengan desain profesional"""
        
        # ========== HEADER ==========
        header_frame = tk.Frame(self.root, bg='#34495e', height=100)
        header_frame.pack(fill='x', pady=(0, 15))
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(
            header_frame,
            text="📊 SISTEM ABSENSI PRO",
            font=('Segoe UI', 24, 'bold'),
            fg='#ffffff',
            bg='#34495e'
        )
        title_label.pack(pady=(15, 5))
        
        subtitle_label = tk.Label(
            header_frame,
            text="Update otomatis ke Excel Template • Support CSV & XLSX",
            font=('Segoe UI', 11),
            fg='#bdc3c7',
            bg='#34495e'
        )
        subtitle_label.pack()
        
        # ========== MAIN CONTAINER ==========
        main_container = tk.Frame(self.root, bg='#ecf0f1')
        main_container.pack(fill='both', expand=True, padx=25, pady=10)
        
        # Left Panel (Input & Settings)
        left_panel = tk.Frame(main_container, bg='#ecf0f1')
        left_panel.pack(side='left', fill='both', expand=True, padx=(0, 10))
        
        # Right Panel (Log & Status)
        right_panel = tk.Frame(main_container, bg='#ecf0f1', width=350)
        right_panel.pack(side='right', fill='both', padx=(10, 0))
        right_panel.pack_propagate(False)
        
        # ========== SECTION 1: FILE INPUT ==========
        self.create_card_section(left_panel, "1️⃣  Pilih File Data Absensi", 0)
        
        file_card = tk.Frame(left_panel, bg='white', relief='flat', bd=0)
        file_card.grid(row=1, column=0, sticky='ew', pady=(0, 15), ipady=15)
        self.add_shadow(file_card)
        
        tk.Label(
            file_card,
            text="File Absensi (CSV/XLSX):",
            font=('Segoe UI', 10, 'bold'),
            bg='white',
            fg='#2c3e50'
        ).grid(row=0, column=0, padx=20, pady=(15, 5), sticky='w')
        
        file_entry_frame = tk.Frame(file_card, bg='white')
        file_entry_frame.grid(row=1, column=0, padx=20, pady=(0, 15), sticky='ew')
        
        file_entry = tk.Entry(
            file_entry_frame,
            textvariable=self.input_file_path,
            font=('Segoe UI', 10),
            state='readonly',
            relief='solid',
            bd=1
        )
        file_entry.pack(side='left', fill='x', expand=True, ipady=8)
        
        browse_btn = tk.Button(
            file_entry_frame,
            text="📁 Browse",
            command=self.browse_input_file,
            font=('Segoe UI', 10, 'bold'),
            bg='#3498db',
            fg='white',
            cursor='hand2',
            relief='flat',
            padx=20,
            pady=8,
            activebackground='#2980b9',
            activeforeground='white'
        )
        browse_btn.pack(side='right', padx=(10, 0))
        self.bind_hover(browse_btn, '#2980b9', '#3498db')
        
        file_card.columnconfigure(0, weight=1)
        
        # ========== SECTION 2: EMPLOYEE SELECTION ==========
        self.create_card_section(left_panel, "2️⃣  Pilih Karyawan & Periode", 2)
        
        selection_card = tk.Frame(left_panel, bg='white', relief='flat', bd=0)
        selection_card.grid(row=3, column=0, sticky='ew', pady=(0, 15), ipady=15)
        self.add_shadow(selection_card)
        
        # Karyawan
        tk.Label(
            selection_card,
            text="Nama Karyawan:",
            font=('Segoe UI', 10, 'bold'),
            bg='white',
            fg='#2c3e50'
        ).grid(row=0, column=0, padx=20, pady=(15, 5), sticky='w')
        
        self.employee_combo = ttk.Combobox(
            selection_card,
            textvariable=self.selected_employee,
            font=('Segoe UI', 10),
            state='readonly'
        )
        self.employee_combo.grid(row=1, column=0, padx=20, pady=(0, 15), sticky='ew')
        
        # Periode Frame
        periode_frame = tk.Frame(selection_card, bg='white')
        periode_frame.grid(row=2, column=0, padx=20, pady=(0, 15), sticky='ew')
        
        # Bulan
        bulan_frame = tk.Frame(periode_frame, bg='white')
        bulan_frame.pack(side='left', fill='x', expand=True, padx=(0, 10))
        
        tk.Label(
            bulan_frame,
            text="Bulan:",
            font=('Segoe UI', 10, 'bold'),
            bg='white',
            fg='#2c3e50'
        ).pack(anchor='w', pady=(0, 5))
        
        self.month_combo = ttk.Combobox(
            bulan_frame,
            textvariable=self.selected_month,
            values=list(range(1, 13)),
            font=('Segoe UI', 10),
            state='readonly',
            width=10
        )
        self.month_combo.pack(fill='x')
        
        # Tahun
        tahun_frame = tk.Frame(periode_frame, bg='white')
        tahun_frame.pack(side='left', fill='x', expand=True)
        
        tk.Label(
            tahun_frame,
            text="Tahun:",
            font=('Segoe UI', 10, 'bold'),
            bg='white',
            fg='#2c3e50'
        ).pack(anchor='w', pady=(0, 5))
        
        self.year_combo = ttk.Combobox(
            tahun_frame,
            textvariable=self.selected_year,
            values=[2024, 2025, 2026, 2027, 2028],
            font=('Segoe UI', 10),
            state='readonly',
            width=10
        )
        self.year_combo.pack(fill='x')
        
        selection_card.columnconfigure(0, weight=1)
        
        # ========== SECTION 3: OUTPUT SETTINGS ==========
        self.create_card_section(left_panel, "3️⃣  File Output Excel", 4)
        
        output_card = tk.Frame(left_panel, bg='white', relief='flat', bd=0)
        output_card.grid(row=5, column=0, sticky='ew', pady=(0, 15), ipady=15)
        self.add_shadow(output_card)
        
        tk.Label(
            output_card,
            text="Nama File Output:",
            font=('Segoe UI', 10, 'bold'),
            bg='white',
            fg='#2c3e50'
        ).grid(row=0, column=0, padx=20, pady=(15, 5), sticky='w')
        
        output_entry = tk.Entry(
            output_card,
            textvariable=self.output_file,
            font=('Segoe UI', 10),
            relief='solid',
            bd=1
        )
        output_entry.grid(row=1, column=0, padx=20, pady=(0, 10), sticky='ew', ipady=8)
        
        tk.Label(
            output_card,
            text="💡 Data akan diupdate pada range: B11-F41",
            font=('Segoe UI', 9, 'italic'),
            bg='white',
            fg='#7f8c8d'
        ).grid(row=2, column=0, padx=20, pady=(0, 15), sticky='w')
        
        output_card.columnconfigure(0, weight=1)
        
        # ========== ACTION BUTTONS ==========
        action_frame = tk.Frame(left_panel, bg='#ecf0f1')
        action_frame.grid(row=6, column=0, sticky='ew', pady=15)
        
        process_btn = tk.Button(
            action_frame,
            text="⚡ PROSES & UPDATE EXCEL",
            command=self.process_and_update,
            font=('Segoe UI', 12, 'bold'),
            bg='#27ae60',
            fg='white',
            cursor='hand2',
            relief='flat',
            padx=30,
            pady=15,
            activebackground='#229954',
            activeforeground='white'
        )
        process_btn.pack(fill='x', pady=(0, 10))
        self.bind_hover(process_btn, '#229954', '#27ae60')
        
        reset_btn = tk.Button(
            action_frame,
            text="🔄 Reset",
            command=self.reset_form,
            font=('Segoe UI', 10),
            bg='#95a5a6',
            fg='white',
            cursor='hand2',
            relief='flat',
            padx=20,
            pady=10,
            activebackground='#7f8c8d',
            activeforeground='white'
        )
        reset_btn.pack(fill='x')
        self.bind_hover(reset_btn, '#7f8c8d', '#95a5a6')
        
        left_panel.columnconfigure(0, weight=1)
        
        # ========== RIGHT PANEL: LOG AREA ==========
        log_label = tk.Label(
            right_panel,
            text="📋 Log Aktivitas",
            font=('Segoe UI', 12, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50'
        )
        log_label.pack(pady=(0, 10))
        
        log_frame = tk.Frame(right_panel, bg='white', relief='flat', bd=0)
        log_frame.pack(fill='both', expand=True)
        self.add_shadow(log_frame)
        
        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            font=('Consolas', 9),
            bg='#2c3e50',
            fg='#ecf0f1',
            relief='flat',
            padx=10,
            pady=10,
            wrap='word',
            state='disabled'
        )
        self.log_text.pack(fill='both', expand=True, padx=2, pady=2)
        
        # Configure log tags
        self.log_text.tag_config('success', foreground='#2ecc71')
        self.log_text.tag_config('error', foreground='#e74c3c')
        self.log_text.tag_config('warning', foreground='#f39c12')
        self.log_text.tag_config('info', foreground='#3498db')
        
        # Initial log
        self.log("Sistem Absensi Pro v2.0 Siap!", 'info')
        self.log("Silakan pilih file data absensi...", 'info')
        
    def create_card_section(self, parent, title, row):
        """Membuat header section dengan style card"""
        label = tk.Label(
            parent,
            text=title,
            font=('Segoe UI', 11, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50',
            anchor='w'
        )
        label.grid(row=row, column=0, sticky='w', pady=(10, 8))
        
    def add_shadow(self, widget):
        """Simulasi shadow untuk widget"""
        widget.configure(relief='flat', bd=0, highlightthickness=1, highlightbackground='#bdc3c7')
        
    def bind_hover(self, button, hover_color, normal_color):
        """Bind hover effect untuk button"""
        button.bind("<Enter>", lambda e: button.config(bg=hover_color))
        button.bind("<Leave>", lambda e: button.config(bg=normal_color))
        
    def log(self, message, tag='info'):
        """Tambahkan pesan ke log area"""
        self.log_text.configure(state='normal')
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.log_text.insert('end', f"[{timestamp}] ", 'info')
        self.log_text.insert('end', f"{message}\n", tag)
        self.log_text.see('end')
        self.log_text.configure(state='disabled')
        self.root.update_idletasks()
        
    def browse_input_file(self):
        """Browse file input (CSV atau XLSX)"""
        file_path = filedialog.askopenfilename(
            title="Pilih File Data Absensi",
            filetypes=[
                ("All Supported", "*.csv *.xlsx"),
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx"),
                ("All files", "*.*")
            ]
        )
        
        if file_path:
            self.input_file_path.set(file_path)
            self.log(f"File dipilih: {os.path.basename(file_path)}", 'success')
            self.load_data_file(file_path)
            
    def load_data_file(self, file_path):
        """Load data dari file CSV atau XLSX"""
        try:
            self.log("Memuat data dari file...", 'info')
            
            # Deteksi ekstensi file
            ext = os.path.splitext(file_path)[1].lower()
            
            if ext == '.csv':
                # Auto-detect separator
                separator = self.detect_csv_separator(file_path)
                self.df_source = pd.read_csv(file_path, sep=separator, dtype=str)
                self.log(f"File CSV berhasil dimuat (separator: '{separator}')", 'success')
                
            elif ext == '.xlsx':
                self.df_source = pd.read_excel(file_path, dtype=str)
                self.log("File XLSX berhasil dimuat", 'success')
            else:
                self.log("Format file tidak didukung!", 'error')
                return
            
            # Bersihkan nama kolom
            self.df_source.columns = self.df_source.columns.astype(str).str.strip().str.lower()
            
            # Ekstrak daftar karyawan
            if 'nama' in self.df_source.columns:
                self.employee_list = self.df_source['nama'].dropna().unique().tolist()
                self.employee_combo['values'] = self.employee_list
                
                if self.employee_list:
                    self.employee_combo.current(0)
                    self.log(f"✓ {len(self.employee_list)} karyawan ditemukan", 'success')
                else:
                    self.log("Tidak ada nama karyawan ditemukan", 'warning')
            else:
                self.log("Kolom 'nama' tidak ditemukan dalam file!", 'error')
                return
            
            # Deteksi periode yang tersedia
            self.detect_available_periods()
            if self.available_periods:
                periods_str = ', '.join([f"{self.get_month_name(m)} {y}" for m, y in self.available_periods])
                self.log(f"✓ Periode tersedia: {periods_str}", 'success')
            else:
                self.log("Tidak dapat mendeteksi periode", 'warning')
                
        except Exception as e:
            self.log(f"Error memuat file: {str(e)}", 'error')
            messagebox.showerror("Error", f"Gagal memuat file:\n{str(e)}")
            
    def detect_csv_separator(self, file_path):
        """Deteksi separator CSV (semicolon atau comma)"""
        try:
            # Coba semicolon
            df = pd.read_csv(file_path, sep=';', dtype=str, nrows=5)
            if len(df.columns) > 3:
                return ';'
            
            # Coba comma
            df = pd.read_csv(file_path, sep=',', dtype=str, nrows=5)
            if len(df.columns) > 3:
                return ','
            
            return ';'  # default
        except:
            return ';'
    
    def detect_available_periods(self):
        """Deteksi bulan dan tahun yang tersedia dalam file"""
        if self.df_source is None:
            return
        
        periods = set()
        abaikan = ['nama', 'nik', 'no', 'nomor']
        date_columns = [col for col in self.df_source.columns if col not in abaikan]
        
        for col in date_columns:
            try:
                date_obj = self.parse_date_column(col)
                
                if date_obj:
                    periods.add((date_obj.month, date_obj.year))
            except:
                continue
        
        self.available_periods = sorted(list(periods))
    
    def parse_date_column(self, col):
        """Parse kolom tanggal, support datetime object (dari XLSX) atau string (dari CSV)"""
        # Jika sudah datetime object (dari XLSX), langsung return
        if isinstance(col, datetime):
            return col
        
        # Jika pandas Timestamp, convert ke datetime
        if hasattr(col, 'to_pydatetime'):
            return col.to_pydatetime()
        
        # Jika string, coba parsing dengan berbagai format
        if isinstance(col, str):
            for fmt in ['%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S']:
                try:
                    return datetime.strptime(col, fmt)
                except ValueError:
                    continue
        
        return None
            
    def get_hari_indonesia(self, date_obj):
        """Dapatkan nama hari dalam Bahasa Indonesia"""
        days = {
            'Monday': 'Senin', 'Tuesday': 'Selasa', 'Wednesday': 'Rabu',
            'Thursday': 'Kamis', 'Friday': 'Jumat', 'Saturday': 'Sabtu',
            'Sunday': 'Minggu'
        }
        return days[date_obj.strftime('%A')]
        
    def is_holiday(self, date_obj):
        """Cek apakah tanggal adalah hari libur nasional"""
        year = str(date_obj.year)
        date_str = date_obj.strftime('%m-%d')
        
        if year in self.INDONESIAN_HOLIDAYS:
            if date_str in self.INDONESIAN_HOLIDAYS[year]:
                return True, self.INDONESIAN_HOLIDAYS[year][date_str]
        
        return False, None
        
    def hitung_durasi(self, masuk_str, pulang_str):
        """Hitung durasi kerja - return integer jam saja"""
        try:
            s1 = str(masuk_str).strip()
            s2 = str(pulang_str).strip()
            fmt1 = '%H:%M:%S' if s1.count(':') >= 2 else '%H:%M'
            fmt2 = '%H:%M:%S' if s2.count(':') >= 2 else '%H:%M'
            t_masuk = datetime.strptime(s1, fmt1)
            t_pulang = datetime.strptime(s2, fmt2)
            selisih = t_pulang - t_masuk
            return int(selisih.total_seconds()) // 3600  # integer jam saja
        except:
            return "-"
            
    def process_attendance_data(self):
        """Proses data absensi untuk karyawan yang dipilih"""
        if self.df_source is None:
            self.log("Data belum dimuat!", 'error')
            return False
            
        employee_name = self.selected_employee.get()
        if not employee_name:
            self.log("Pilih karyawan terlebih dahulu!", 'error')
            return False
            
        month = self.selected_month.get()
        year = self.selected_year.get()
        
        self.log(f"Memproses data untuk: {employee_name}", 'info')
        self.log(f"Periode: {self.get_month_name(month)} {year}", 'info')
        
        try:
            # Cari data karyawan
            pegawai = self.df_source[self.df_source['nama'].str.contains(
                employee_name, case=False, na=False)]
            
            if pegawai.empty:
                self.log(f"Karyawan '{employee_name}' tidak ditemukan!", 'error')
                return False
            
            # Kolom tanggal (selain nama, nik, no, nomor)
            abaikan = ['nama', 'nik', 'no', 'nomor']
            date_columns = [col for col in self.df_source.columns if col not in abaikan]
            
            output_data = []
            
            for col in date_columns:
                # Parse header tanggal (support datetime object dari XLSX dan string dari CSV)
                try:
                    date_obj = self.parse_date_column(col)
                    
                    if date_obj is None:
                        continue
                        
                except:
                    continue
                
                # Filter bulan dan tahun
                if date_obj.month != month or date_obj.year != year:
                    continue
                
                hari = self.get_hari_indonesia(date_obj)
                is_libur, nama_libur = self.is_holiday(date_obj)
                
                # Ambil nilai absensi
                raw_value = pegawai.iloc[0][col]
                
                jam_masuk = "-"
                jam_pulang = "-"
                durasi = "-"
                keterangan = "-"
                
                val_str = str(raw_value).strip()
                
                # Logika pengisian data
                if pd.isna(raw_value) or val_str == "" or val_str.lower() == "nan":
                    if is_libur:
                        keterangan = f"Libur - {nama_libur}"
                    elif hari in ['Sabtu', 'Minggu']:
                        keterangan = hari  # Tampilkan nama hari (Sabtu/Minggu)
                    else:
                        keterangan = "Tidak Hadir / Tanpa Keterangan"
                else:
                    # Format: "HH:MM:SS - HH:MM:SS"
                    parts = val_str.split('-')
                    
                    if len(parts) >= 2:
                        raw_masuk = parts[0].strip()
                        raw_pulang = parts[1].strip()
                        
                        if ':' in raw_masuk and ':' in raw_pulang:
                            durasi = self.hitung_durasi(raw_masuk, raw_pulang)
                        
                        # Format HH:MM (tanpa detik)
                        jam_masuk = ':'.join(raw_masuk.split(':')[:2]) if ':' in raw_masuk else raw_masuk
                        jam_pulang = ':'.join(raw_pulang.split(':')[:2]) if ':' in raw_pulang else raw_pulang
                        
                        if is_libur:
                            keterangan = f"Hadir di Hari Libur ({nama_libur})"
                        else:
                            keterangan = ""  # Kosongkan untuk status hadir normal
                            
                    elif len(parts) == 1 and parts[0] != "":
                        raw_masuk = parts[0].strip()
                        jam_masuk = ':'.join(raw_masuk.split(':')[:2]) if ':' in raw_masuk else raw_masuk
                        keterangan = "Absen Tidak Lengkap"
                
                output_data.append({
                    'tanggal': date_obj,
                    'tanggal_str': date_obj.strftime('%d/%m/%Y'),
                    'jam_masuk': jam_masuk,
                    'jam_pulang': jam_pulang,
                    'durasi': durasi,
                    'keterangan': keterangan
                })
            
            if not output_data:
                self.log("❌ Tidak ada data untuk periode yang dipilih!", 'error')
                
                # Tampilkan periode yang tersedia
                if self.available_periods:
                    periods_str = ', '.join([f"{self.get_month_name(m)} {y}" for m, y in self.available_periods])
                    self.log(f"ℹ️  Periode tersedia: {periods_str}", 'info')
                    messagebox.showwarning(
                        "Data Tidak Ditemukan",
                        f"Tidak ada data untuk {self.get_month_name(month)} {year}!\n\n"
                        f"Periode yang tersedia dalam file:\n{periods_str}\n\n"
                        f"Silakan pilih periode yang sesuai."
                    )
                else:
                    messagebox.showwarning(
                        "Data Tidak Ditemukan",
                        f"Tidak ada data untuk {self.get_month_name(month)} {year}!"
                    )
                return False
            
            # Sort by tanggal
            output_data.sort(key=lambda x: x['tanggal'])
            
            self.processed_data = output_data
            self.log(f"✓ {len(output_data)} hari berhasil diproses", 'success')
            return True
            
        except Exception as e:
            self.log(f"Error memproses data: {str(e)}", 'error')
            return False
            
    def update_excel_file(self):
        """Update file Excel hasil-akhir.xlsx dengan data yang diproses"""
        output_file = self.output_file.get()
        
        if not output_file:
            output_file = "hasil-akhir.xlsx"
            
        if not output_file.endswith('.xlsx'):
            output_file += '.xlsx'
        
        try:
            self.log(f"Mempersiapkan file: {output_file}", 'info')
            
            # Path template
            template_path = "contoh/hasil-akhir.xlsx"
            
            # Cek apakah template ada
            if os.path.exists(template_path):
                self.log(f"Menggunakan template: {template_path}", 'info')
                # Load dengan rich_text=True agar CellRichText (bold sebagian) bisa ditulis
                wb = load_workbook(template_path, rich_text=True)
                ws = wb.active
                # Bersihkan gambar yang sudah terbaca dari template
                ws._images.clear()
            elif os.path.exists(output_file):
                self.log("Template tidak ada, menggunakan file existing...", 'warning')
                wb = load_workbook(output_file)
                ws = wb.active
            else:
                self.log("Template & file tidak ada, membuat file baru...", 'warning')
                wb = Workbook()
                ws = wb.active
                ws.title = "Rekap Absensi"
                
                # Create header template jika file baru
                self.create_excel_template(ws)
            
            # Update informasi karyawan dan periode di header
            try:
                month_name = self.get_month_name(self.selected_month.get())
                month = self.selected_month.get()  # diperlukan untuk datetime()
                year = self.selected_year.get()
                periode_text = f"{month_name} {year}"

                # Update nama karyawan (F7) - jaga font asli
                orig_font_f7 = ws['F7'].font
                ws['F7'].value = self.selected_employee.get()
                ws['F7'].font = Font(
                    name=orig_font_f7.name or 'Calibri',
                    size=orig_font_f7.size or 12,
                    bold=orig_font_f7.bold,
                    italic=orig_font_f7.italic,
                    color=orig_font_f7.color
                )
                self.log("✓ Nama karyawan diupdate di F7", 'success')

                # Update G5 dengan datetime agar number_format mmm-yy bekerja
                ws['G5'].value = datetime(year, month, 1)
                ws['G5'].number_format = 'mmm-yy'
                ws['G5'].font = Font(name='Calibri', size=18, color=Color(theme=4))
                self.log(f"✓ Periode diupdate di G5: {periode_text}", 'success')

                # Pastikan D5 tetap bold '-' (separator) dengan warna biru (theme=4)
                ws['D5'].value = '-'
                ws['D5'].font = Font(name='Calibri', size=18, bold=True, color=Color(theme=4))

                # Pastikan B44 & E44 tetap bold
                ws['B44'].font = Font(name='Calibri', size=11, bold=True)
                ws['E44'].font = Font(name='Calibri', size=11, bold=True)

                # Update nama karyawan di C45 (kolom tanda tangan tenaga ahli)
                ws['C45'].value = self.selected_employee.get()
                self.log(f"✓ Nama karyawan diupdate di C45", 'success')

                # Bold sebagian teks header organisasi di C1:
                # 3 baris pertama bold, 3 baris bawah normal
                bold_part = (
                    'PEMERINTAH KABUPATEN BADUNG\n'
                    'DINAS KOMUNIKASI DAN INFORMATIKA\n'
                    'PUSAT PEMERINTAHAN MANGUPRAJA MANDALA\n'
                )
                normal_part = (
                    'Jln Raya Sempidi, Mengwi \u2013 Kabupaten Badung (80351)\n'
                    'Telp. (0361) 419888 Fax (0361) 419888\n'
                    'Website : www.diskominfo.badungkab.go.id'
                )
                ws['C1'].value = CellRichText(
                    TextBlock(InlineFont(rFont='Calibri', sz=12, b=True), bold_part),
                    TextBlock(InlineFont(rFont='Calibri', sz=12, b=False), normal_part)
                )
                self.log('✓ Header organisasi C1 diformat (bold sebagian)', 'success')

                # Header tabel row 10 - hapus background, tidak bold
                no_fill = PatternFill(fill_type=None)
                for col in ['B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}10'].fill = no_fill

                # C49 & F49: tanggal hari ini dengan format 3-Jan-2026
                today = datetime.now()
                ws['C49'].value = today
                ws['C49'].number_format = r'[$-409]d\-mmm\-yyyy;@'
                ws['F49'].value = today
                ws['F49'].number_format = r'[$-409]d\-mmm\-yyyy;@'
                self.log(f"✓ Tanggal hari ini diisi di C49 & F49", 'success')

            except Exception as e:
                self.log(f"Info header mungkin perlu disesuaikan manual: {str(e)}", 'warning')
            
            # Update data ke range B11:F41
            start_row = 11
            max_rows = 31  # B11 sampai B41
            
            self.log("Menulis data ke Excel (B11:F41)...", 'info')
            
            for idx, data in enumerate(self.processed_data):
                if idx >= max_rows:
                    self.log(f"Maksimal {max_rows} baris tercapai, data selanjutnya dilewati", 'warning')
                    break
                
                row = start_row + idx
                
                # B: Tanggal - format d-Mmm-yy sesuai template
                ws[f'B{row}'] = data['tanggal']
                ws[f'B{row}'].number_format = r'[$-409]d\-mmm\-yy;@'
                
                # C: Jam Masuk - format HH:MM
                ws[f'C{row}'] = data['jam_masuk']
                ws[f'C{row}'].number_format = 'h:mm;@'
                
                # D: Jam Keluar/Pulang - format HH:MM
                ws[f'D{row}'] = data['jam_pulang']
                ws[f'D{row}'].number_format = 'h:mm;@'
                
                # E: Durasi (integer jam)
                ws[f'E{row}'] = data['durasi']
                if isinstance(data['durasi'], int):
                    ws[f'E{row}'].number_format = '#,##0'
                
                # F: Keterangan
                ws[f'F{row}'] = data['keterangan']
            
            # Clear sisa baris jika data kurang dari 31
            for idx in range(len(self.processed_data), max_rows):
                row = start_row + idx
                ws[f'B{row}'] = ""
                ws[f'C{row}'] = ""
                ws[f'D{row}'] = ""
                ws[f'E{row}'] = ""
                ws[f'F{row}'] = ""
            
            # Sisipkan logo sebelum simpan
            logo_path = "contoh/logo-badung.png"
            if os.path.exists(logo_path):
                try:
                    logo_img = XLImage(logo_path)
                    logo_img.width = 121
                    logo_img.height = 116
                    # Gunakan TwoCellAnchor dengan koordinat EMU persis dari template asli
                    # agar posisi logo identik (rata tengah di area B1:B4)
                    anchor = TwoCellAnchor(editAs='oneCell')
                    anchor._from = AnchorMarker(col=1, colOff=249464, row=0, rowOff=68036)
                    anchor.to   = AnchorMarker(col=1, colOff=1401989, row=3, rowOff=86270)
                    logo_img.anchor = anchor
                    ws.add_image(logo_img)
                    self.log("✓ Logo berhasil ditambahkan (posisi rata tengah)", 'success')
                except Exception as e:
                    self.log(f"Gagal menambahkan logo: {str(e)}", 'warning')
            else:
                self.log("ℹ️  Logo tidak ditemukan (contoh/logo-badung.png)", 'warning')

            # Tentukan path simpan: prioritas ke Windows Downloads
            win_dl = self.get_windows_downloads_path()
            if win_dl:
                save_path = os.path.join(win_dl, os.path.basename(output_file))
                self.log(f"✓ Target simpan: Windows Downloads ({save_path})", 'info')
            else:
                save_path = output_file
                self.log("ℹ️  Windows Downloads tidak ditemukan, simpan lokal", 'warning')

            # Simpan file
            wb.save(save_path)
            self.log(f"✓ File berhasil disimpan: {save_path}", 'success')
            
            messagebox.showinfo(
                "Berhasil!",
                f"File disimpan ke:\n{save_path}\n\n"
                f"Range: B11:F{start_row + len(self.processed_data) - 1}\n"
                f"Total: {len(self.processed_data)} hari"
            )
            
            return True
            
        except Exception as e:
            self.log(f"Error menyimpan Excel: {str(e)}", 'error')
            messagebox.showerror("Error", f"Gagal menyimpan file Excel:\n{str(e)}")
            return False
            
    def create_excel_template(self, ws):
        """Buat template Excel sederhana jika file baru"""
        # Header utama
        ws['A1'] = 'REKAP ABSENSI KARYAWAN'
        ws['A1'].font = Font(size=16, bold=True)
        
        ws['A3'] = 'Nama'
        ws['B3'] = self.selected_employee.get()
        
        ws['A4'] = 'Periode'
        ws['B4'] = f"{self.get_month_name(self.selected_month.get())} {self.selected_year.get()}"
        
        # Header tabel (row 10)
        headers = ['No', 'Tanggal', 'Jam Masuk', 'Jam Keluar', 'Durasi', 'Keterangan']
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=10, column=idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Nomor urut (A11:A41)
        for i in range(1, 32):
            ws[f'A{10+i}'] = i

    def get_windows_downloads_path(self):
        """Deteksi folder Windows Downloads via WSL mount /mnt/c/Users/"""
        try:
            users_dir = '/mnt/c/Users'
            if os.path.exists(users_dir):
                skip = {'public', 'default', 'default user', 'all users'}
                users = [
                    d for d in os.listdir(users_dir)
                    if os.path.isdir(os.path.join(users_dir, d))
                    and d.lower() not in skip
                ]
                if users:
                    dl = os.path.join(users_dir, users[0], 'Downloads')
                    if os.path.exists(dl):
                        return dl
        except Exception:
            pass
        return None
            
    def process_and_update(self):
        """Proses data dan update file Excel"""
        # Validasi
        if not self.input_file_path.get():
            messagebox.showwarning("Peringatan", "Pilih file data absensi terlebih dahulu!")
            return
            
        if not self.selected_employee.get():
            messagebox.showwarning("Peringatan", "Pilih karyawan terlebih dahulu!")
            return
        
        self.log("="*50, 'info')
        self.log("MEMULAI PROSES...", 'info')
        
        # Proses data
        if self.process_attendance_data():
            # Update Excel
            self.update_excel_file()
            self.log("PROSES SELESAI!", 'success')
            self.log("="*50, 'info')
        else:
            self.log("PROSES GAGAL!", 'error')
            self.log("="*50, 'info')
            
    def reset_form(self):
        """Reset form ke kondisi awal"""
        self.input_file_path.set("")
        self.selected_employee.set("")
        self.df_source = None
        self.employee_list = []
        self.processed_data = None
        self.available_periods = []
        self.employee_combo['values'] = []
        
        self.log("Form direset", 'info')
        
    @staticmethod
    def get_month_name(month_number):
        """Dapatkan nama bulan dalam Bahasa Indonesia"""
        months = {
            1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
            5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
            9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
        }
        return months.get(month_number, '')


def main():
    """Main function"""
    root = tk.Tk()
    app = SistemAbsensiPro(root)
    
    # Center window
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    root.mainloop()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n⚠️  Program dihentikan oleh user")
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        input("\nTekan Enter untuk keluar...")
